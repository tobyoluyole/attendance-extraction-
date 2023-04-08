'''import bluetooth
import xlsxwriter
import openpyxl'''

'''nearby_devices = bluetooth.discover_devices(duration=30, flush_cache=True, lookup_names=False, lookup_class=False)

def read_details():
    for addr in nearby_devices:
        return(addr)'''


'''workbook = xlsxwriter.Workbook('student_details.xlsx')
worksheet = workbook.add_worksheet()
#worksheet.set_column('A:A', 20)
# Rows and columns are zero indexed.
row = 0
column = 0

content =(read_details())
for item in content :
    worksheet.write(row=item+1, column=1, item=content)
    #row += 1


workbook.close()'''
import bluetooth
import openpyxl

# Scan for nearby Bluetooth devices
devices = bluetooth.discover_devices(duration=30, flush_cache=True, lookup_names=False, lookup_class=False)

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.column_dimensions['A'].width = 20

# Write Bluetooth MAC addresses to worksheet
for i, addr in enumerate(devices):
    worksheet.cell(row=i+1, column=1, value=addr)

# Save Excel workbook
workbook.save("bluetooth_addresses.xlsx")

print("Found {} Bluetooth MAC addresses. Saved to 'bluetooth_addresses.xlsx'".format(len(devices)))

