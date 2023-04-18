import bluetooth
import openpyxl

# Scan for nearby Bluetooth devices
devices = bluetooth.discover_devices(duration=30, flush_cache=True, lookup_names=False, lookup_class=False)

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.column_dimensions['A'].width = 20

# Write Bluetooth MAC addresses to worksheet
for i, addr in enumerate(devices):
    worksheet.cell(row=i+1, column=1, value=addr)

# Save Excel workbook
workbook.save("bluetooth_addresses.xlsx")

print("Found {} Bluetooth MAC addresses. Saved to 'bluetooth_addresses.xlsx'".format(len(devices)))

import pandas as pd
import sqlite3

# read the excel file into a pandas DataFrame
excel_file = pd.read_excel('bluetooth_addresses.xlsx', header=None, names=['Mac_addresses'])

# connect to the database
conn = sqlite3.connect('Student_Information.db')
cursor = conn.cursor()

# retrieve the names of the students whose addresses are present in both the excel file and the database
result = cursor.execute("""
    SELECT Full_name FROM Students_details
    WHERE Mac_addresses IN (%s)
""" % ','.join(['?' for i in range(len(excel_file))]), tuple(excel_file['Mac_addresses'].tolist())).fetchall()

# print the names of the students
with open('output.txt', 'w') as f:
    for row in result:
        f.write(row[0] + '\n')
# close the database connection
cursor.close()
conn.close()

import openpyxl

# Load the Attendance excel file
wb = openpyxl.load_workbook('Attendance.xlsx')
sheet = wb.active

# Get the column index to start marking attendance
last_column = sheet.max_column
start_column = last_column + 1

# Open the output file with the names
with open('output.txt', 'r') as f:
    names = f.read().splitlines()

# Loop through the names and mark attendance in the Attendance excel file
for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value
    if name in names:
        sheet.cell(row=row, column=start_column).value = True
    else:
        sheet.cell(row=row, column=start_column).value = False

# Save the Attendance excel file
wb.save('Attendance.xlsx')



import os
import smtplib
import datetime 
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Set up the email addresses and password
sender_email = "isktobi@gmail.com"
sender_password = "gusunnkswqqswdbn"
receiver_email = "tobioluyole@gmail.com, oluyolevincent@gmail.com"

# Create a message object instance
message = MIMEMultipart()

# Set the message parameters
message['From'] = sender_email
message['To'] = receiver_email
message['Subject'] = "Attendance Report for Today"

# Add the body of the email (including today's date)
today_date = datetime.date.today().strftime('%Y-%m-%d')
message.attach(MIMEText("Please find attached the attendance report for today ({})".format(today_date)))

# Attach the Excel file
filename = "Attendance.xlsx"
with open(filename, "rb") as attachment:
    part = MIMEApplication(
        attachment.read(),
        Name=os.path.basename(filename)
    )
    part['Content-Disposition'] = 'attachment; filename="{}"'.format(os.path.basename(filename))
    message.attach(part)

# Connect to the SMTP server and send the email
with smtplib.SMTP("smtp.gmail.com", 587) as server:
    server.starttls()
    server.login(sender_email, sender_password)
    text = message.as_string()
    server.sendmail(sender_email, receiver_email.split(','), text)

print("Attendance report sent sucessfully")




